#!/usr/bin/env python3
"""
◈조치완료◈ 편지함 메일 정보 추출 → 엑셀 저장 스크립트 (v3)

흐름:
  1. 자동 로그인 (auto_login.py 재사용)
  2. ◈조치완료◈ 편지함 이동
  3. 각 메일 클릭
  4. 첨부파일 옆 미리보기 버튼(evt-rol="read-nested-pop") 클릭
  5. 팝업에서 보낸사람/받는사람/날짜/제목/첨부파일명 추출
  6. 엑셀 저장
"""

import os
import sys
import os
os.environ['PYTHONUTF8'] = '1'
os.environ['PYTHONIOENCODING'] = 'utf-8'
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import time
import re
from datetime import datetime

os.environ["PYTHONUTF8"] = "1"
os.environ["PYTHONIOENCODING"] = "utf-8"

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] pip install openpyxl")
    sys.exit(1)

try:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException
    from selenium.webdriver.common.keys import Keys
except ImportError:
    print("[!] pip install selenium")
    sys.exit(1)

# ─────────────────────────────────────────────
# 임포트 경로 설정 (EXE 실행 대응)
# ─────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    # PyInstaller 임시 폴더(sys._MEIPASS)를 가장 먼저 탐색하도록 설정
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

if base_path not in sys.path:
    sys.path.insert(0, base_path)

# 모듈 임포트 (PyInstaller 분석기가 명확히 인식하도록 직접 임포트)
from auto_login import load_config, login_shinhan_mail


# ─────────────────────────────────────────────
# ◈조치완료◈ 편지함 이동
# ─────────────────────────────────────────────
def navigate_to_malmail_folder(driver):
    print("\n[*] ◈조치완료◈ 편지함으로 이동 중...")

    # iframe 탐색
    for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.frame(iframe)
            if _click_malmail_element(driver):
                driver.switch_to.default_content()
                try:
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr[id]")))
                except TimeoutException:
                    pass
                time.sleep(0.5)
                return True
            driver.switch_to.default_content()
        except Exception:
            driver.switch_to.default_content()

    # 기본 컨텍스트
    driver.switch_to.default_content()
    if _click_malmail_element(driver):
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr[id]")))
        except TimeoutException:
            pass
        time.sleep(0.5)
        return True

    print("[!] ◈조치완료◈ 찾기 실패. 폴더 목록:")
    try:
        for el in driver.find_elements(By.CSS_SELECTOR, "li[evt-rol='folder'] a"):
            print(f"  - '{el.text.strip()}'")
    except Exception:
        pass
    return False


def _click_malmail_element(driver):
    for kw in ["◈조치완료◈", "조치완료"]:
        try:
            for el in driver.find_elements(By.XPATH, f"//*[contains(text(),'{kw}')]"):
                if el.is_displayed():
                    driver.execute_script("arguments[0].click();", el)
                    print(f"[+] '{kw}' 클릭 완료")
                    return True
        except Exception:
            continue
    return False


def _format_date(date_str):
    """
    다양한 날짜 문자열을 YYYY-MM-DD HH:MM:SS 형식으로 변환합니다.
    예: '2026/03/04 화요일 4:07:11' → '2026-03-04 04:07:11'
    """
    if not date_str:
        return date_str
    # 한국어 요일 제거 (월/화/수/목/금/토/일 + 요일)
    cleaned = re.sub(r'[월화수목금토일]요일\s*', '', date_str)
    # AM/PM 오전/오후 제거
    cleaned = re.sub(r'(오전|오후|AM|PM)\s*', '', cleaned, flags=re.IGNORECASE)
    cleaned = cleaned.strip()
    # YYYY/MM/DD H:MM:SS 또는 YYYY-MM-DD H:MM:SS 패턴
    m = re.match(r'(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})\s+(\d{1,2}):(\d{2}):(\d{2})', cleaned)
    if m:
        y, mo, d, hh, mm, ss = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d} {int(hh):02d}:{mm}:{ss}"
    # 시:분만 있는 경우
    m2 = re.match(r'(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})\s+(\d{1,2}):(\d{2})', cleaned)
    if m2:
        y, mo, d, hh, mm = m2.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d} {int(hh):02d}:{mm}:00"
    # YY-MM-DD HH:MM 형태 (신한메일 목록 포맷)
    m3 = re.match(r'(\d{2})[/\-.](\d{1,2})[/\-.](\d{1,2})\s+(\d{1,2}):(\d{2})', cleaned)
    if m3:
        y, mo, d, hh, mm = m3.groups()
        y_int = int(y)
        y_full = 2000 + y_int if y_int < 50 else 1900 + y_int
        return f"{y_full}-{int(mo):02d}-{int(d):02d} {int(hh):02d}:{mm}:00"
    return date_str  # 변환 실패 시 원본 반환


def _find_row_by_id(driver, mail_id):
    """
    JavaScript querySelector로 특수문자(&, =, + 등) 포함 ID를 안전하게 탐색합니다.
    """
    try:
        # JavaScript CSS.escape 활용
        js = """
        var mid = arguments[0];
        try {
            return document.querySelector('tr[id="' + mid + '"]');
        } catch(e) {
            // fallback: 모든 tr[id] 탐색
            var rows = document.querySelectorAll('table tbody tr[id]');
            for (var i = 0; i < rows.length; i++) {
                if (rows[i].id === mid) return rows[i];
            }
            return null;
        }
        """
        el = driver.execute_script(js, mail_id)
        return el
    except Exception:
        return None


# 조치완료함 URL (go_back_to_list에서 사용)
MALMAIL_URL = ""

# ─────────────────────────────────────────────
# 메일 ID 수집 (페이지네이션)
# ─────────────────────────────────────────────
EXCLUDE_ID_RE = re.compile(
    r'^(allSelectTr|dateDesc_|dateAsc_|toolbar_|pageNavi)',
    re.IGNORECASE
)

def collect_all_mail_ids(driver, target_limit=None):
    print("\n[*] 메일 목록 수집 중...")
    _set_page_size(driver, "80")

    all_ids = []
    page = 1

    while True:
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr[id]")))
        except TimeoutException:
            pass
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr[id]")
        mail_ids = [
            r.get_attribute("id") for r in rows
            if r.get_attribute("id") and not EXCLUDE_ID_RE.match(r.get_attribute("id"))
        ]
        print(f"[*] 페이지 {page}: {len(mail_ids)}개 메일")
        if not mail_ids and page == 1:
            print("[-] 메일 없음")
            return []
        for mid in mail_ids:
            if mid not in all_ids:
                all_ids.append(mid)
                if target_limit and len(all_ids) >= target_limit:
                    print(f"[*] 목표 추출 건수({target_limit}건) 달성. 목록 수집을 조기 종료합니다.")
                    return all_ids[:target_limit]
                    
        try:
            nb = driver.find_element(By.CSS_SELECTOR, "#pageNaviWrap a.next.paginate_button")
            if "paginate_button_disabled" in (nb.get_attribute("class") or ""):
                break
            driver.execute_script("arguments[0].click();", nb)
            page += 1
            # 페이지 전환 후 메일 행이 새로 렌더링될 때까지 동적 대기
            try:
                WebDriverWait(driver, 5).until(
                    EC.staleness_of(rows[0])
                )
            except Exception:
                pass
            time.sleep(0.5)
        except Exception:
            break

    print(f"[*] 총 {len(all_ids)}개 수집 완료")
    return all_ids


def _set_page_size(driver, value="80"):
    try:
        from selenium.webdriver.support.ui import Select
        sel = Select(driver.find_element(By.CSS_SELECTOR, "#toolbar_list_pagebase"))
        if sel.first_selected_option.get_attribute("value") != value:
            sel.select_by_value(value)
            print(f"[*] 페이지당 {value}건으로 변경")
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr[id]")))
            except TimeoutException:
                pass
            time.sleep(1)
    except Exception as e:
        print(f"[!] 페이지당 건수 변경 실패: {e}")


# ─────────────────────────────────────────────
# 메일 클릭 → 미리보기 버튼 클릭 → 팝업 정보 추출
# ─────────────────────────────────────────────
def open_mail_and_extract_info(driver, mail_id):
    """
    메일 행 클릭 → 첨부파일 옆 미리보기 버튼(evt-rol=read-nested-pop) 클릭
    → 팝업에서 보낸사람/받는사람/날짜/제목/첨부파일명 추출
    """
    info = {
        "신고 일시": "", "메일제목": "", "보낸사람": "", "받는사람": "",
        "보낸날짜": "", "첨부파일명": "",
        "그룹사": "", "부서": "", "신고자": "",
    }

    # ── 메일 행 찾기 (JS querySelector로 특수문자 포함 ID 안전 처리) ──
    row = _find_row_by_id(driver, mail_id)
    if row is None:
        print(f"  [!] 메일 행 못 찾음: {mail_id}")
        return None

    # ── 신고 일시 (목록에서 추출) ──
    try:
        dates_in_list = row.find_elements(By.CSS_SELECTOR, "td.date, span.date, div.date")
        if dates_in_list:
            info["신고 일시"] = _format_date((dates_in_list[0].text or "").strip())
        if not info["신고 일시"]:
            for td in row.find_elements(By.TAG_NAME, "td"):
                txt = (td.text or "").strip()
                if re.search(r'\d{2}[:\.]\d{2}', txt) or "요일" in txt:
                    info["신고 일시"] = _format_date(txt)
                    if info["신고 일시"]: break
    except Exception as e:
        print(f"  [!] 신고 일시 파싱 예외 발생: {e}")

    # 목록에서 제목 미리 수집 (클릭 전 실행 필수)
    try:
        # 1차 시도: 특정 셀렉터 위주
        subject_found = False
        subject_targets = row.find_elements(By.CSS_SELECTOR, "a.subject, .mail_subject, td.subject a, span.subject")
        if not subject_targets:
            subject_targets = row.find_elements(By.TAG_NAME, "a")

        for lk in subject_targets:
            txt = (lk.text or "").strip()
            # UI 텍스트 제외 필터 강화
            if txt and len(txt) > 2 and not any(k in txt for k in ["첨부파일", "파일열기", "미리보기"]):
                raw_title = txt[:120]
                prefix_pattern = re.compile(r'^\[(신고메일|Report email|举报邮件)\]\s*', re.IGNORECASE)
                cleaned_title = prefix_pattern.sub('', raw_title)
                
                info["메일제목"] = cleaned_title
                print(f"  [+] 원본 제목 수집 완료: {cleaned_title}")
                subject_found = True
                break
        
        # 2차 시도: 셀렉터 실패 시 td 전체 텍스트에서 추출
        if not subject_found:
            for td in row.find_elements(By.TAG_NAME, "td"):
                txt = (td.text or "").strip()
                # 공백이나 숫자만 있는 경우 제외, 너무 짧은 경우 제외
                if txt and len(txt) > 5 and not any(k in txt for k in ["첨부파일", "요일", "2026", "2025", "오전", "오후"]):
                    # 신고메일 패턴이 포함된 경우 우선순위
                    if "[신고메일]" in txt or "[Report email]" in txt or "[举报邮件]" in txt:
                        raw_title = txt[:120]
                        prefix_pattern = re.compile(r'^\[(신고메일|Report email|举报邮件)\]\s*', re.IGNORECASE)
                        info["메일제목"] = prefix_pattern.sub('', raw_title)
                        print(f"  [+] 목록 텍스트에서 제목 추출: {info['메일제목']}")
                        subject_found = True
                        break
    except Exception as e:
        print(f"  [!] 목록 제목 수집 예외: {e}")

    # ── 1. 메일 클릭 ──
    try:
        # 본문 로드 대기 - a 태그나 onclick 속성이 있는 요소 클릭
        clicked = False
        click_targets = row.find_elements(By.CSS_SELECTOR, "a.subject, .mail_subject, td.subject a, span.subject")
        if not click_targets:
            click_targets = row.find_elements(By.TAG_NAME, "td")
        
        for target in click_targets:
            if target.is_displayed() and target.text.strip():
                driver.execute_script("arguments[0].click();", target)
                clicked = True
                break
                
        if not clicked:
            driver.execute_script("arguments[0].click();", row)
            
        print(f"  [본문 진입 시도]")
        
        # 본문 렌더링 확인 (미리보기 버튼이나 메일 헤더가 나타날 때까지 대기)
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "span[evt-rol='read-nested-pop'], .btn_fn4, .mail_read_wrap, .mail_header"))
            )
        except TimeoutException:
            pass

    except Exception as e:
        print(f"  [!] 메일 클릭 단계 접근 실패: {e}")

    # 본문 URL 로깅 (어차피 SPA이므로 현재 URL)
    print(f"  [본문 URL] {driver.current_url}")

    # (기존: 첨부파일명으로 제목을 덮어쓰던 로직 제거 완료 - 원본 제목 유지)

    # ── #readContentMessageWrap > iframe > body#contentBox > .reportPhishing 테이블 파싱 ──
    KW_MAP = [
        (["소속회사", "회사명", "그룹사", "소속", "affiliated company", "affiliatedcompany", "所属公司"], "그룹사"),
        (["부서명", "부서", "department", "部门"],                                                    "부서"),
        (["신고자", "성명", "이름", "담당자", "reporter's name", "reporter", "reportersname", "举报人姓名"], "신고자"),
    ]
    try:
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.ID, "readContentMessageWrap"))
        )
        # iframe 내부 콘텐츠 로드를 동적으로 대기 (고정 sleep 제거)
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#readContentMessageWrap iframe"))
            )
        except Exception:
            pass
    except Exception:
        pass
    try:
        def _parse_report_table(ctx, require_class=True):
            """reportPhishing 테이블 또는 일반 테이블에서 th→td 쌍을 파싱해 dict 반환."""
            result = {}
            # 1차: .reportPhishing 클래스 내부 테이블 파싱
            containers = ctx.find_elements(By.CSS_SELECTOR, ".reportPhishing")
            if not containers and not require_class:
                # 2차: .reportPhishing 가 없으면 모든 테이블 대상으로 파싱
                containers = ctx.find_elements(By.CSS_SELECTOR, "table")
            for container in containers:
                for row in container.find_elements(By.TAG_NAME, "tr"):
                    ths = row.find_elements(By.TAG_NAME, "th")
                    tds = row.find_elements(By.TAG_NAME, "td")
                    if not ths or not tds:
                        continue
                    lbl = (ths[0].text or "").strip().replace(" ", "").lower()
                    val = (tds[0].text or "").strip()
                    for kws, field in KW_MAP:
                        if result.get(field): continue
                        if any(kw in lbl for kw in kws) and val:
                            result[field] = val[:100]
            return result

        parsed = {}

        # 1단계: 현재 컨텍스트 직접 파싱 (.reportPhishing 우선)
        if not any(parsed.values()):
            parsed = _parse_report_table(driver)
        # 1-1단계: .reportPhishing 없으면 일반 테이블도 시도
        if not any(parsed.values()):
            parsed = _parse_report_table(driver, require_class=False)

        # 2단계: 레벨1 iframe 탐색
        if not any(parsed.values()):
            lv1_iframes = driver.find_elements(By.CSS_SELECTOR,
                                               "#readContentMessageWrap iframe, iframe")
            for iframe1 in lv1_iframes:
                try:
                    driver.switch_to.frame(iframe1)
                    try:
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located(
                                (By.CSS_SELECTOR, ".reportPhishing, table, body"))
                        )
                    except Exception:
                        pass
                    parsed = _parse_report_table(driver)
                    if not any(parsed.values()):
                        parsed = _parse_report_table(driver, require_class=False)
                    if any(parsed.values()):
                        driver.switch_to.default_content()
                        break

                    # 3단계: 레벨2 중첩 iframe 탐색 (parent_frame으로 복귀)
                    lv2_iframes = driver.find_elements(By.TAG_NAME, "iframe")
                    for iframe2 in lv2_iframes:
                        try:
                            driver.switch_to.frame(iframe2)
                            try:
                                WebDriverWait(driver, 3).until(
                                    EC.presence_of_element_located(
                                        (By.CSS_SELECTOR, ".reportPhishing, table, body"))
                                )
                            except Exception:
                                pass
                            parsed = _parse_report_table(driver)
                            if not any(parsed.values()):
                                parsed = _parse_report_table(driver, require_class=False)
                            driver.switch_to.parent_frame()  # 레벨1로 복귀
                            if any(parsed.values()):
                                break
                        except Exception:
                            try: driver.switch_to.default_content()
                            except Exception: pass
                            break

                    driver.switch_to.default_content()
                    if any(parsed.values()):
                        break
                except Exception:
                    try: driver.switch_to.default_content()
                    except Exception: pass

        for field, val in parsed.items():
            if val:
                info[field] = val
                print(f"  [+] {field}: {val}")

        if not any(parsed.values()):
            print("  [*] reportPhishing 섹션에서 데이터를 찾지 못함")
    except Exception as e:
        print(f"  [!] reportPhishing 파싱 실패: {e}")
        try: driver.switch_to.default_content()
        except Exception: pass

    # ── [Fallback] 테이블 파싱 실패 시 본문 텍스트에서 정규식으로 추출 ──
    if not all([info.get("그룹사"), info.get("부서"), info.get("신고자")]):
        try:
            driver.switch_to.default_content()
            full_text = ""
            # 현재 페이지 + iframe 내부 텍스트 수집
            try:
                full_text = driver.find_element(By.TAG_NAME, "body").text or ""
            except Exception:
                pass
            for iframe in driver.find_elements(By.CSS_SELECTOR, "#readContentMessageWrap iframe, iframe"):
                try:
                    driver.switch_to.frame(iframe)
                    full_text += "\n" + (driver.find_element(By.TAG_NAME, "body").text or "")
                    # 레벨2 iframe도 수집
                    for iframe2 in driver.find_elements(By.TAG_NAME, "iframe"):
                        try:
                            driver.switch_to.frame(iframe2)
                            full_text += "\n" + (driver.find_element(By.TAG_NAME, "body").text or "")
                            driver.switch_to.parent_frame()
                        except Exception:
                            try: driver.switch_to.parent_frame()
                            except: pass
                    driver.switch_to.default_content()
                except Exception:
                    try: driver.switch_to.default_content()
                    except: pass

            if full_text:
                # 정규식 패턴: "라벨 : 값" 또는 "라벨: 값" 형태
                FALLBACK_PATTERNS = [
                    ("그룹사", re.compile(
                        r'(?:소속회사|회사명|그룹사|소속|affiliated\s*company|所属公司)\s*[:：]\s*(.+)',
                        re.IGNORECASE)),
                    ("부서", re.compile(
                        r'(?:부서명|부서|department|部门)\s*[:：]\s*(.+)',
                        re.IGNORECASE)),
                    ("신고자", re.compile(
                        r'(?:신고자|성명|이름|담당자|reporter\'?s?\s*name|举报人姓名)\s*[:：]\s*(.+)',
                        re.IGNORECASE)),
                ]
                for field, pattern in FALLBACK_PATTERNS:
                    if info.get(field):
                        continue
                    m = pattern.search(full_text)
                    if m:
                        val = m.group(1).strip()[:100]
                        if val:
                            info[field] = val
                            print(f"  [+] {field} (fallback): {val}")
        except Exception as e:
            print(f"  [!] fallback 텍스트 파싱 실패: {e}")
            try: driver.switch_to.default_content()
            except Exception: pass

    # ── 첨부파일 옆 미리보기 버튼 탐색 ──
    # 셀렉터: span[evt-rol="read-nested-pop"] 또는 .btn_fn4
    preview_clicked = False
    preview_selectors = [
        'span[evt-rol="read-nested-pop"]',
        '.btn_fn4[evt-rol="read-nested-pop"]',
        '.btn_fn4',
        '[evt-rol="read-nested-pop"]',
    ]

    for sel in preview_selectors:
        try:
            btns = driver.find_elements(By.CSS_SELECTOR, sel)
            for btn in btns:
                if btn.is_displayed():
                    txt = (btn.text or "").strip()
                    print(f"  [+] 미리보기 버튼 발견: '{txt}' (셀렉터: {sel})")
                    driver.execute_script("arguments[0].click();", btn)
                    preview_clicked = True
                    # 고정된 시간이 아닌 작은 sleep 후 WebDriverWait으로 대체 가능하지만, 애니메이션 위해 1초 대기
                    time.sleep(1)
                    break
            if preview_clicked:
                break
        except Exception:
            continue

    if not preview_clicked:
        print("  [!] 미리보기 버튼을 찾지 못했습니다. 본문에서 직접 추출합니다.")
        # 미리보기 버튼 없는 경우 본문에서 직접 추출
        info = _extract_from_mail_body(driver, info)
        return info

    # ── 팝업 처리 ──
    info = _extract_from_popup(driver, info)
    return info


# ─────────────────────────────────────────────
# 팝업에서 정보 추출
# ─────────────────────────────────────────────
def _extract_from_popup(driver, info):
    """
    미리보기 팝업(mailPopup.do 새 창)에서 메일 정보 추출.
    [실제 구조 - 디버그로 확인]
      - 닫기: .btn_minor_s (텍스트 '닫기') 또는 .btn_layer_x (X)
      - 발신: span.name (첫 번째)
      - 수신: span.name (두 번째)
      - 날짜: span.date
      - 제목: span.subject
    """
    original_window = driver.current_window_handle
    time.sleep(2)

    # ── 새 창(mailPopup.do) 처리 ──
    all_windows = driver.window_handles
    if len(all_windows) > 1:
        new_win = [w for w in all_windows if w != original_window][-1]
        driver.switch_to.window(new_win)
        print(f"  [*] 팝업 창 전환: {driver.current_url}")
        
        # 새 창의 DOM이 렌더링되기를 짧게 기다림
        try:
            WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        except Exception:
            pass
        time.sleep(0.5)

        # ── 데이터 추출 (mailPopup.do 전용 셀렉터) ──
        try:
            # 제목: span.subject (이미 EML 파일명으로 제목이 설정된 경우 덮어쓰지 않음)
            # 제목: span.subject (이미 EML 파일명으로 제목이 설정된 경우 덮어쓰지 않음)
            if not info["메일제목"]:
                subjs = driver.find_elements(By.CSS_SELECTOR, "span.subject")
                for subj in subjs:
                    subj_txt = (subj.text or "").strip()
                    # 팝업 UI 텍스트(첨부파일 개수, 용량 등)를 제목으로 가져오지 않도록 필터 강화
                    if subj_txt and len(subj_txt) > 2:
                        excluded_keywords = ["첨부파일", " 개", "KB", "MB", "GB", "파일열기", "다운로드"]
                        if not any(k in subj_txt for k in excluded_keywords):
                            info["메일제목"] = subj_txt
                            break

            # 날짜: span.date → YYYY-MM-DD HH:MM:SS 포맷 변환
            dates = driver.find_elements(By.CSS_SELECTOR, "span.date")
            if dates:
                info["보낸날짜"] = _format_date((dates[0].text or "").strip())

            # 보낸사람 / 받는사람: span.name (순서대로 첫번째=발신, 두번째=수신)
            names = driver.find_elements(By.CSS_SELECTOR, "span.name")
            if len(names) >= 1:
                info["보낸사람"] = (names[0].text or "").strip()
            if len(names) >= 2:
                info["받는사람"] = (names[1].text or "").strip()

            # tr 구조 fallback (span.title - span.name 쌍)
            if not (info["보낸사람"] and info["받는사람"]):
                for row in driver.find_elements(By.CSS_SELECTOR, "tr"):
                    t_els = row.find_elements(By.CSS_SELECTOR, "span.title")
                    v_els = row.find_elements(By.CSS_SELECTOR, "span.name, span.date")
                    if t_els and v_els:
                        lbl = (t_els[0].text or "").strip()
                        val = (v_els[0].text or "").strip()
                        if any(k in lbl for k in ["발", "보낸", "from", "From"]):
                            if not info["보낸사람"]: info["보낸사람"] = val
                        elif any(k in lbl for k in ["수", "받", "to", "To"]):
                            if not info["받는사람"]: info["받는사람"] = val
                        elif any(k in lbl for k in ["날", "date", "Date"]):
                            if not info["보낸날짜"]: info["보낸날짜"] = val

        except Exception as e:
            print(f"  [!] 팝업 데이터 추출 오류: {e}")

        # ── 첨부파일명 추출: 파일명 (크기KB) 패턴에서 파일명 부분만 추출 ──
        try:
            attach_names = []
            # 전체 텍스트에서 파일명(size) 패턴 적용
            body_text = driver.find_element(By.TAG_NAME, "body").text or ""
            # "filename.ext (number,numberKB)" 또는 "filename.ext (numberKB)" 패턴
            for m in re.finditer(
                r'([^\n\r]+?\.[a-zA-Z]{2,6})\s*\(\d[\d,\.]*\s*KB\)',
                body_text, re.IGNORECASE
            ):
                fname = m.group(1).strip()
                # 제외 키워드 (실제 파일명이 아니얰혼되는 것 제외)
                if fname and not any(k in fname for k in ["단위", "MB", "GB", "달러", "USD"]):

                    attach_names.append(fname)

            # CSS 셀렉터 방식 보완
            if not attach_names:
                for el in driver.find_elements(By.CSS_SELECTOR,
                        "span.ic_file_s, [class*='file_name'], [class*='attach'] span.name"):
                    txt = (el.text or "").strip()
                    if txt and '.' in txt and len(txt) < 200:
                        attach_names.append(txt)

            if attach_names:
                info["첨부파일명"] = ", ".join(attach_names)
                print(f"  [+] 첨부: {info['첨부파일명']}")
        except Exception as e:
            print(f"  [!] 첨부파일명 추출 실패: {e}")

        # ── 닫기 버튼 클릭 (실제 구조 기반) ──
        closed = False

        # 1단계: .btn_minor_s '닫기' 버튼
        if not closed:
            try:
                for cb in driver.find_elements(By.CSS_SELECTOR, ".btn_minor_s"):
                    if "닫" in (cb.text or "") and cb.is_displayed():
                        driver.execute_script("arguments[0].click();", cb)
                        time.sleep(1.5)
                        if new_win not in driver.window_handles:
                            closed = True
                            print("  [*] '.btn_minor_s 닫기' 버튼으로 팝업 종료")
                        break
            except Exception:
                pass

        # 2단계: .btn_layer_x (X 아이콘)
        if not closed:
            try:
                for xb in driver.find_elements(By.CSS_SELECTOR, ".btn_layer_x"):
                    if xb.is_displayed():
                        driver.execute_script("arguments[0].click();", xb)
                        time.sleep(1.5)
                        if new_win not in driver.window_handles:
                            closed = True
                            print("  [*] '.btn_layer_x' X버튼으로 팝업 종료")
                        break
            except Exception:
                pass

        # 3단계: driver.close() Selenium fallback
        if not closed:
            try:
                driver.close()
                time.sleep(1)
                closed = True
                print("  [*] driver.close()로 팝업 종료")
            except Exception as e:
                print(f"  [!] 팝업 종료 실패: {e}")


        # 원래 창으로 복귀
        try:
            driver.switch_to.window(original_window)
            print("  [*] 원래 창으로 복귀")
        except Exception:
            try:
                driver.switch_to.window(driver.window_handles[0])
            except Exception:
                pass
        return info

    # ── HTML 오버레이 팝업 처리 ──
    # 신한 메일 미리보기 팝업은 보통 특정 id나 class를 가진 layer
    popup_selectors = [
        # 신한 메일 nested mail 뷰 팝업
        ".layer_wrap",         ".mail_layer",
        "[class*='nested']",   "[class*='popup']",
        "[class*='layer']",    "[class*='modal']",
        "[role='dialog']",
        ".mail_area_read",
        # 일반적인 팝업
        ".swal2-popup",        ".modal.show",
    ]

    popup_el = None
    for ps in popup_selectors:
        try:
            for el in driver.find_elements(By.CSS_SELECTOR, ps):
                if el.is_displayed():
                    popup_el = el
                    print(f"  [*] 오버레이 팝업 감지: {ps}")
                    break
        except Exception:
            continue
        if popup_el:
            break

    if popup_el:
        # 팝업 내 iframe도 확인
        inner_iframes = []
        try:
            inner_iframes = popup_el.find_elements(By.TAG_NAME, "iframe")
        except Exception:
            pass

        if inner_iframes:
            for iframe in inner_iframes:
                try:
                    driver.switch_to.frame(iframe)
                    result = _parse_header_fields(driver)
                    driver.switch_to.default_content()
                    if any(result.values()):
                        _merge_info(info, result)
                        break
                except Exception:
                    driver.switch_to.default_content()
        else:
            result = _parse_header_fields_from_el(popup_el, driver)
            _merge_info(info, result)

        # 팝업 닫기
        _close_popup(driver, popup_el)
    else:
        # 팝업을 못 찾은 경우 → JavaScript로 팝업 요소 재탐색
        print("  [*] CSS 팝업 감지 실패. JS로 가시 요소 재탐색...")
        try:
            # 화면에 새로 나타난 요소들을 모두 가져와 헤더 파싱 시도
            result = _parse_header_fields(driver)
            _merge_info(info, result)
        except Exception:
            pass

    return info


def _parse_popup_content(driver, info):
    """팝업 창 또는 특정 컨텍스트에서 메일 헤더를 파싱합니다."""
    # iframe 내부 확인
    for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.frame(iframe)
            result = _parse_header_fields(driver)
            driver.switch_to.default_content()
            if any(result.values()):
                _merge_info(info, result)
                return info
        except Exception:
            driver.switch_to.default_content()

    # 기본 컨텍스트
    driver.switch_to.default_content()
    result = _parse_header_fields(driver)
    _merge_info(info, result)
    return info


# ─────────────────────────────────────────────
# 헤더 파싱 공통 함수
# ─────────────────────────────────────────────
def _parse_header_fields(driver):
    """현재 컨텍스트에서 메일 헤더 필드를 추출합니다."""
    result = {"메일제목": "", "보낸사람": "", "받는사람": "", "보낸날짜": "", "첨부파일명": ""}
    _extract_from_table(driver, result)
    _extract_from_dl(driver, result)
    _extract_by_selectors(driver, result)
    if not all([result["보낸사람"], result["받는사람"], result["보낸날짜"]]):
        _extract_from_text_body(driver, result)
    if not result["첨부파일명"]:
        result["첨부파일명"] = _extract_attachments(driver)
    return result


def _parse_header_fields_from_el(element, driver):
    """HTML 요소 내부에서 헤더 필드 추출."""
    result = {"메일제목": "", "보낸사람": "", "받는사람": "", "보낸날짜": "", "첨부파일명": ""}
    try:
        # 요소 내 th-td
        for row in element.find_elements(By.CSS_SELECTOR, "tr"):
            ths = row.find_elements(By.TAG_NAME, "th")
            tds = row.find_elements(By.TAG_NAME, "td")
            if ths and tds:
                # 라벨 비교용 텍스트 (공백, 콜론 제거)
                label_for_cmp = (ths[0].text or "").strip().replace(" ","").replace(":","").lower()
                # 실제 데이터 값 (콜론 유지)
                value = (tds[0].text or "").strip()
                if label_for_cmp and value:
                    _assign_field(label_for_cmp, value, result)
    except Exception:
        pass
    # 텍스트 전체 파싱
    try:
        lines = [(element.text or "").split("\n")]
        lines = [l.strip() for l in lines[0] if l.strip()]
        _parse_text_lines(lines, result)
    except Exception:
        pass
    return result


def _extract_from_mail_body(driver, info):
    """미리보기 버튼이 없을 때 본문 페이지에서 직접 추출."""
    for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.frame(iframe)
            result = _parse_header_fields(driver)
            driver.switch_to.default_content()
            if any(result.values()):
                _merge_info(info, result)
                return info
        except Exception:
            driver.switch_to.default_content()
    driver.switch_to.default_content()
    result = _parse_header_fields(driver)
    _merge_info(info, result)
    return info


def _extract_from_table(driver, result):
    try:
        for row in driver.find_elements(By.CSS_SELECTOR, "tr"):
            ths = row.find_elements(By.TAG_NAME, "th")
            tds = row.find_elements(By.TAG_NAME, "td")
            if not ths or not tds:
                continue
            # 라벨 비교용 (공백, 콜론 제거)
            lbl_for_cmp = (ths[0].text or "").strip().replace(" ","").replace(":","").lower()
            # 데이터 값 (콜론 유지)
            val = (tds[0].text or "").strip()
            if lbl_for_cmp and val:
                _assign_field(lbl_for_cmp, val, result)
    except Exception:
        pass


def _extract_from_dl(driver, result):
    try:
        dts = driver.find_elements(By.TAG_NAME, "dt")
        dds = driver.find_elements(By.TAG_NAME, "dd")
        for dt, dd in zip(dts, dds):
            lbl = (dt.text or "").strip().replace(" ","").lower()
            val = (dd.text or "").strip()
            if lbl and val:
                _assign_field(lbl, val, result)
    except Exception:
        pass


def _extract_by_selectors(driver, result):
    """CSS 셀렉터로 직접 추출. 신한 메일 뷰어 패턴 포함."""
    MAP = {
        "메일제목": [
            ".mail_subject", ".view_subject", "#mailSubject",
            "[class*='subject']", "h3", "h2",
        ],
        "보낸사람": [
            ".mail_from em", ".from_addr", "#mailFrom",
            "[class*='from'] em", "[class*='sender']",
            "span[class*='from']", "td[class*='from']",
        ],
        "받는사람": [
            ".mail_to em", ".to_addr", "#mailTo",
            "[class*='rcpt']", "[class*='to'] em",
        ],
        "보낸날짜": [
            ".mail_date", "#mailDate", "[class*='date'] em",
            "[class*='date'] span", ".send_date",
        ],
    }
    for field, sels in MAP.items():
        if result[field]:
            continue
        for sel in sels:
            try:
                for el in driver.find_elements(By.CSS_SELECTOR, sel):
                    txt = (el.text or "").strip()
                    if txt and 1 < len(txt) < 500:
                        result[field] = txt
                        break
                if result[field]:
                    break
            except Exception:
                continue


def _extract_from_text_body(driver, result):
    """본문 텍스트에서 라벨:값 패턴 매칭."""
    try:
        lines = [(driver.find_element(By.TAG_NAME, "body").text or "").split("\n")]
        lines = [l.strip() for l in lines[0] if l.strip()]
        _parse_text_lines(lines, result)
    except Exception:
        pass


def _parse_text_lines(lines, result):
    KW = {
        "메일제목":   ["제목", "subject"],
        "보낸사람":   ["보낸사람", "from", "발신자", "발신"],
        "받는사람":   ["받는사람", "to", "수신자", "수신", "받는"],
        "보낸날짜":   ["날짜", "date", "일시", "보낸날짜"],
    }
    for i, line in enumerate(lines):
        ls = line.lower().replace(" ", "")
        for field, kws in KW.items():
            if result[field]:
                continue
            for kw in kws:
                if kw in ls:
                    if ":" in line:
                        parts = line.split(":", 1)
                        if len(parts) == 2 and parts[1].strip():
                            result[field] = parts[1].strip()[:500]
                    elif i + 1 < len(lines):
                        nxt = lines[i + 1]
                        nxt_ls = nxt.lower().replace(" ", "")
                        if not any(k in nxt_ls for k in ["제목","보낸사람","받는사람","날짜","from","to","date","subject"]):
                            result[field] = nxt[:500]
                    break


def _extract_attachments(driver):
    """첨부파일 목록 추출."""
    names = []
    for sel in [
        "#attachListWrap li",
        ".attach_list li",
        '[evt-rol="download-attach"]',
        '.name[evt-rol*="attach"]',
    ]:
        try:
            for item in driver.find_elements(By.CSS_SELECTOR, sel):
                txt = (item.text or "").strip()
                if txt:
                    fname = re.split(r'\s*[\(\[（【]', txt)[0].strip()
                    if fname and fname not in names:
                        names.append(fname)
            if names:
                break
        except Exception:
            continue
    return ", ".join(names) if names else ""


def _assign_field(label_lower, value, result):
    if any(k in label_lower for k in ["제목", "subject"]):
        if not result["메일제목"]:
            result["메일제목"] = value
    elif any(k in label_lower for k in ["보낸", "from", "발신"]):
        if not result["보낸사람"]:
            result["보낸사람"] = value
    elif any(k in label_lower for k in ["받는", "to", "수신", "rcpt"]):
        if not result["받는사람"]:
            result["받는사람"] = value
    elif any(k in label_lower for k in ["날짜", "date", "일시"]):
        if not result["보낸날짜"]:
            result["보낸날짜"] = value


def _merge_info(info, result):
    for k in info:
        if not info[k] and result.get(k):
            info[k] = result[k]


def _close_popup(driver, popup_el=None):
    """팝업 닫기."""
    close_kws = ["닫기", "close", "취소", "×", "✕", "X"]
    # 팝업 내 닫기 버튼
    if popup_el:
        try:
            for btn in popup_el.find_elements(By.CSS_SELECTOR, "button,a,[class*='close'],[class*='btn']"):
                txt = (btn.text or "").strip()
                cls = (btn.get_attribute("class") or "")
                if any(k.lower() in txt.lower() or k in cls for k in close_kws):
                    driver.execute_script("arguments[0].click();", btn)
                    print("  [*] 팝업 닫기")
                    time.sleep(1)
                    return
        except Exception:
            pass
    # 전체 페이지에서 닫기
    try:
        for btn in driver.find_elements(By.CSS_SELECTOR, "button,[class*='close'],[class*='닫기']"):
            try:
                txt = (btn.text or "").strip()
                cls = (btn.get_attribute("class") or "")
                if btn.is_displayed() and any(k.lower() in txt.lower() or k.lower() in cls.lower() for k in close_kws):
                    driver.execute_script("arguments[0].click();", btn)
                    print("  [*] 팝업 닫기(페이지)")
                    time.sleep(1)
                    return
            except Exception:
                continue
    except Exception:
        pass
    # ESC
    try:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        print("  [*] ESC로 팝업 닫기")
        time.sleep(1)
    except Exception:
        pass


# ─────────────────────────────────────────────
# 목록 복귀
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# 목록 복귀
# ─────────────────────────────────────────────
def go_back_to_list(driver):
    """
    메일 본문에서 ◈조치완료◈ 목록으로 복귀합니다.
    기존처럼 사이드바의 폴더를 클릭하면 1페이지로 무조건 초기화되므로,
    메일 뷰어 상단의 '[목록]' 버튼을 누르거나 브라우저 뒤로가기를 최우선으로 사용합니다.
    """
    global MALMAIL_URL

    navigated = False
    
    # 1차: 상단/하단의 '목록' 버튼 (class btn_fn3 등) 탐색 및 클릭
    # 신한 메일의 목록 버튼 셀렉터 후보
    list_selectors = [
        "a.btn_fn3[evt-rol='go-list']", 
        "button[evt-rol='go-list']",
        ".btnWrap .btn_fn3",
        "a[title='목록']",
        "button[title='목록']"
    ]
    
    for sel in list_selectors:
        try:
            for btn in driver.find_elements(By.CSS_SELECTOR, sel):
                txt = (btn.text or "").strip()
                if btn.is_displayed() and ("목록" in txt or "List" in txt or txt == ""):
                    driver.execute_script("arguments[0].click();", btn)
                    print("  [*] 상단/하단 '[목록]' 버튼 클릭으로 복귀 (상태 유지)")
                    navigated = True
                    time.sleep(1.5)
                    break
            if navigated:
                break
        except Exception:
            continue

    # 2차: 목록 버튼이 없거나 실패 시 브라우저 기본 뒤로가기 (상태 유지 기능 기대)
    if not navigated:
        try:
            driver.back()
            print("  [*] 브라우저 Back() 으로 복귀 (상태 유지)")
            navigated = True
            time.sleep(1.5)
        except Exception:
            pass
            
    # 3차: 모두 실패 시 URL 직접 이동 (최후의 수단, 1페이지로 초기화됨)
    if not navigated:
        if MALMAIL_URL:
            print("  [!] 목록 복귀/뒤로가기 모두 실패 → URL 직접 이동 (초기화)")
            driver.get(MALMAIL_URL)
            time.sleep(2)


# ─────────────────────────────────────────────
# 엑셀 저장
# ─────────────────────────────────────────────
def save_to_excel(records, output_dir=None):
    if not records:
        print("[!] 저장할 데이터가 없습니다.")
        return None

    if output_dir is None:
        if getattr(sys, 'frozen', False):
            output_dir = os.path.dirname(sys.executable)
        else:
            output_dir = os.path.dirname(os.path.abspath(__file__))

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"malmail_info_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "조치완료 목록"

    hdr_font  = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="000000")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["No", "신고 일시", "메일 유형", "메일 유입 시간", "발신자", "수신자", "메일제목", "첨부파일명", "유형", "그룹사", "부서", "신고자"]
    col_widths = [5, 22, 15, 22, 32, 32, 50, 35, 15, 25, 25, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 25

    dat_font  = Font(name="맑은 고딕", size=10)
    dat_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    even_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")

    for ri, rec in enumerate(records, 2):
        row_data = [
            ri - 1,
            rec.get("신고 일시", ""),
            rec.get("메일 유형", ""),
            rec.get("메일 유입 시간", ""),
            rec.get("발신자", ""),
            rec.get("수신자", ""),
            rec.get("메일제목", ""),
            rec.get("첨부파일명", ""),
            rec.get("유형", ""),
            rec.get("그룹사", ""),
            rec.get("부서", ""),
            rec.get("신고자", ""),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dat_font; c.alignment = dat_align; c.border = border
            if ri % 2 == 0:
                c.fill = even_fill
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    try:
        wb.save(path)
        print(f"\n[SUCCESS] 엑셀 저장: {path}")
        print(f"[*] 총 {len(records)}건")
        return path
    except Exception as e:
        print(f"[!] 엑셀 저장 실패: {e}")
        return None


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  ◈조치완료◈ 편지함 정보 추출 → 엑셀 저장 (v3)")
    print("  첨부파일 옆 미리보기 버튼(read-nested-pop) 활용")
    print("=" * 60)
    
    # +++ 1. 추출 건수 입력받기 +++
    if len(sys.argv) > 1:
        limit_input = sys.argv[1].strip()
        print(f"  [*] CLI 입력 감지: {limit_input}")
    else:
        limit_input = input("  몇 건의 이메일 정보를 가져오시겠습니까? (전체 가져오려면 그냥 Enter): ").strip()
        
    target_limit = None
    if limit_input.isdigit() and int(limit_input) > 0:
        target_limit = int(limit_input)
        print(f"  [*] 목표 추출 건수: {target_limit}건")
    else:
        print("  [*] 목표 추출 건수: 전체")
    print("=" * 60)

    config = load_config()
    driver = login_shinhan_mail(config)
    if not driver:
        print("[FAIL] 로그인 실패")
        sys.exit(1)
        
    # Selenium 글로벌 implicit wait 비활성화 (지연의 핵심 원인)
    driver.implicitly_wait(0)
    
    print("[SUCCESS] 로그인 완료!\n")

    records = []
    try:
        if not navigate_to_malmail_folder(driver):
            print("[!] ◈조치완료◈ 폴더 이동 실패")
            input("Enter 후 종료...")
            driver.quit()
            sys.exit(1)

        time.sleep(3)
        # 조치완료함 URL 저장 (목록 복귀 시 재사용)
        global MALMAIL_URL
        MALMAIL_URL = driver.current_url
        print(f"[*] 조치완료함 URL: {MALMAIL_URL}")

        # 제한 변수를 넘겨서 지정 수량까지만 수집
        mail_ids = collect_all_mail_ids(driver, target_limit)
        
        # 만약을 대비해 수집된 배열 재슬라이싱
        if target_limit and len(mail_ids) > target_limit:
            mail_ids = mail_ids[:target_limit]
            
        if not mail_ids:
            print("[!] 메일 없음")
            driver.quit()
            sys.exit(0)

        # 첫 페이지 복귀 (이미 1페이지면 생략)
        try:
            fb = driver.find_element(By.CSS_SELECTOR, "#pageNaviWrap a.first.paginate_button")
            if fb.is_displayed() and "paginate_button_disabled" not in (fb.get_attribute("class") or ""):
                driver.execute_script("arguments[0].click();", fb)
                time.sleep(1.5)
        except Exception:
            pass

        print(f"\n[*] 총 {len(mail_ids)}개 메일 처리 시작\n")

        for idx, mail_id in enumerate(mail_ids):
            print(f"\n{'─'*55}")
            print(f"[*] 메일 {idx+1}/{len(mail_ids)}  (ID: {mail_id})")

            # 메일 행이 현재 페이지에 없으면 다음 페이지를 탐색하거나 조치완료함으로 이동
            if _find_row_by_id(driver, mail_id) is None:
                print("  [*] 현재 페이지에 행 없음 → 다음 페이지 이동 탐색 시도")
                found = False
                
                # 1. 현재 화면에서 '다음' 버튼 누르며 찾아보기
                for _ in range(30):
                    try:
                        nb = driver.find_element(By.CSS_SELECTOR, "#pageNaviWrap a.next.paginate_button")
                        if "paginate_button_disabled" in (nb.get_attribute("class") or ""):
                            break
                        driver.execute_script("arguments[0].click();", nb)
                        time.sleep(3)
                        if _find_row_by_id(driver, mail_id):
                            found = True
                            print("  [*] 다음 페이지에서 행 발견 완료")
                            break
                    except Exception:
                        break
                        
                # 2. 그래도 못찾으면 조치완료 메뉴부터 새로고침하여 1페이지부터 다시 탐색 (안전망)
                if not found:
                    print("  [*] 탐색 실패, 조치완료함 초기화 후 재탐색 시작")
                    driver.get(MALMAIL_URL)
                    time.sleep(5)
                    _set_page_size(driver, "80")
                    time.sleep(2)
                    for _ in range(30):
                        if _find_row_by_id(driver, mail_id):
                            found = True
                            print("  [*] 재로딩 후 행 발견 완료")
                            break
                        try:
                            nb = driver.find_element(By.CSS_SELECTOR, "#pageNaviWrap a.next.paginate_button")
                            if "paginate_button_disabled" in (nb.get_attribute("class") or ""):
                                break
                            driver.execute_script("arguments[0].click();", nb)
                            time.sleep(3)
                        except Exception:
                            break

            info = open_mail_and_extract_info(driver, mail_id)
            if info:
                # 발신자 텍스트에서 < > 사이 내용만 파싱
                sender_raw = info.get("보낸사람", "")
                match = re.search(r'<([^>]+)>', sender_raw)
                sender = match.group(1).strip() if match else sender_raw

                # 엑셀 저장용 컬럼명 및 순서 매핑
                final_info = {
                    "신고 일시": info.get("신고 일시", ""),
                    "메일 유형": "악성 메일",
                    "메일 유입 시간": info.get("보낸날짜", ""),
                    "발신자": sender,
                    "수신자": info.get("받는사람", ""),
                    "메일제목": info.get("메일제목", ""),
                    "첨부파일명": info.get("첨부파일명", ""),
                    "유형": "유입 신고",
                    "그룹사": info.get("그룹사", ""),
                    "부서": info.get("부서", ""),
                    "신고자": info.get("신고자", ""),
                }

                records.append(final_info)
                print(f"  제목:     {final_info['메일제목'][:50]}")
                print(f"  발신자:   {final_info['발신자']}")
                print(f"  수신자:   {final_info['수신자']}")
                print(f"  유입시간: {final_info['메일 유입 시간']}")
                print(f"  신고일시: {final_info['신고 일시']}")
                print(f"  그룹사:   {final_info['그룹사'] or '(없음)'}")
                print(f"  부서:     {final_info['부서'] or '(없음)'}")
                print(f"  신고자:   {final_info['신고자'] or '(없음)'}")
                print(f"  첨부:     {final_info.get('첨부파일명') if final_info.get('첨부파일명') else '(없음)'}")

            go_back_to_list(driver)
            time.sleep(0.5)
            # 페이지 사이즈는 초기 구동시나 복귀실패(직접이동) 시퀀스에서만 맞춤
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        save_to_excel(records, base_dir)

        print("\n" + "=" * 60)
        print(f"  완료: {len(mail_ids)}건 조회 / {len(records)}건 수집")
        print("=" * 60)

    except Exception as e:
        print(f"\n[!] 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        if records:
            print(f"[*] 중간 데이터 {len(records)}건 저장...")
            save_to_excel(records)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print("[*] 브라우저 종료")


if __name__ == "__main__":
    main()

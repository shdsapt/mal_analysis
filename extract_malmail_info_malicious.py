#!/usr/bin/env python3
"""
◈악성메일◈ 편지함 메일 정보 추출 → 엑셀 저장 스크립트 (v3.5 - 최신 패치 적용본)

수정 내역:
  - 중국어 키워드 지원 (所属公司, 部门, 举报人姓名)
  - 제목 파싱 로직 강화 (목록 2중 수집, 팝업 UI 텍스트 필터링)
  - 인코딩 이슈 해결 (콜론 유지, UTF-8 강제 옵션 대응)
  - 파싱 안정성 강화 (WebDriverWait, Fallback 정규식 로직)
  - 중복 수집 방지 (mail_ids 기반 정밀 처리)
"""

import os
import sys
import time
import re
from datetime import datetime

# 인코딩 강제 설정 (EXE 실행 및 다른 PC 환경 대응)
os.environ['PYTHONUTF8'] = '1'
os.environ['PYTHONIOENCODING'] = 'utf-8'
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] pip install openpyxl")
    sys.exit(1)

try:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException
    from selenium.webdriver.common.keys import Keys
except ImportError:
    print("[!] pip install selenium")
    sys.exit(1)

# ─────────────────────────────────────────────
# 임포트 경로 설정 (EXE 실행 대응)
# ─────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

if base_path not in sys.path:
    sys.path.insert(0, base_path)

from auto_login import load_config, login_shinhan_mail


# ─────────────────────────────────────────────
# ◈악성메일◈ 편지함 이동 (유지)
# ─────────────────────────────────────────────
def navigate_to_malmail_folder(driver):
    print("\n[*] ◈악성메일◈ 편지함으로 이동 중...")
    
    # iframe 탐색
    for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.frame(iframe)
            if _click_malmail_element(driver):
                driver.switch_to.default_content()
                try:
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr[id]")))
                except: pass
                time.sleep(0.5)
                return True
            driver.switch_to.default_content()
        except:
            driver.switch_to.default_content()

    driver.switch_to.default_content()
    if _click_malmail_element(driver):
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr[id]")))
        except: pass
        time.sleep(0.5)
        return True
    return False

def _click_malmail_element(driver):
    for kw in ["◈악성메일◈", "악성메일"]:
        try:
            for el in driver.find_elements(By.XPATH, f"//*[contains(text(),'{kw}')]"):
                if el.is_displayed():
                    driver.execute_script("arguments[0].click();", el)
                    print(f"[+] '{kw}' 클릭 완료")
                    return True
        except: continue
    return False


# ─────────────────────────────────────────────
# 공통 유틸리티 및 파싱 함수
# ─────────────────────────────────────────────
def _format_date(date_str):
    if not date_str: return date_str
    cleaned = re.sub(r'[월화수목금토일]요일\s*', '', date_str)
    cleaned = re.sub(r'(오전|오후|AM|PM)\s*', '', cleaned, flags=re.IGNORECASE).strip()
    
    m = re.match(r'(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})\s+(\d{1,2}):(\d{2}):(\d{2})', cleaned)
    if m:
        y, mo, d, hh, mm, ss = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d} {int(hh):02d}:{mm}:{ss}"
    
    m2 = re.match(r'(\d{2})[/\-.](\d{1,2})[/\-.](\d{1,2})\s+(\d{1,2}):(\d{2})', cleaned)
    if m2:
        y, mo, d, hh, mm = m2.groups()
        y_int = int(y)
        y_full = 2000 + y_int if y_int < 50 else 1900 + y_int
        return f"{y_full}-{int(mo):02d}-{int(d):02d} {int(hh):02d}:{mm}:00"
    return date_str

def _find_row_by_id(driver, mail_id):
    try:
        js = """
        var mid = arguments[0];
        try {
            return document.querySelector('tr[id="' + mid + '"]');
        } catch(e) {
            var rows = document.querySelectorAll('table tbody tr[id]');
            for (var i = 0; i < rows.length; i++) {
                if (rows[i].id === mid) return rows[i];
            }
            return null;
        }
        """
        return driver.execute_script(js, mail_id)
    except: return None

MALMAIL_URL = ""
EXCLUDE_ID_RE = re.compile(r'^(allSelectTr|dateDesc_|dateAsc_|toolbar_|pageNavi)', re.IGNORECASE)

def collect_all_mail_ids(driver, target_limit=None):
    print("\n[*] 메일 목록 수집 중...")
    _set_page_size(driver, "80")
    all_ids = []
    page = 1
    while True:
        try: WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr[id]")))
        except: pass
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr[id]")
        mail_ids = [r.get_attribute("id") for r in rows if r.get_attribute("id") and not EXCLUDE_ID_RE.match(r.get_attribute("id"))]
        print(f"[*] 페이지 {page}: {len(mail_ids)}개 메일")
        if not mail_ids and page == 1: return []
        for mid in mail_ids:
            if mid not in all_ids:
                all_ids.append(mid)
                if target_limit and len(all_ids) >= target_limit: return all_ids[:target_limit]
        try:
            nb = driver.find_element(By.CSS_SELECTOR, "#pageNaviWrap a.next.paginate_button")
            if "paginate_button_disabled" in (nb.get_attribute("class") or ""): break
            driver.execute_script("arguments[0].click();", nb)
            page += 1
            time.sleep(1)
        except: break
    print(f"[*] 총 {len(all_ids)}개 수집 완료")
    return all_ids

def _set_page_size(driver, value="80"):
    try:
        from selenium.webdriver.support.ui import Select
        sel = Select(driver.find_element(By.CSS_SELECTOR, "#toolbar_list_pagebase"))
        if sel.first_selected_option.get_attribute("value") != value:
            sel.select_by_value(value)
            time.sleep(1.5)
    except: pass


# ─────────────────────────────────────────────
# 메일 정보 추출 엔진
# ─────────────────────────────────────────────
def open_mail_and_extract_info(driver, mail_id):
    info = {"신고 일시": "", "메일제목": "", "보낸사람": "", "받는사람": "", "보낸날짜": "", "첨부파일명": "", "그룹사": "", "부서": "", "신고자": ""}
    row = _find_row_by_id(driver, mail_id)
    if not row: return None

    # 신고 일시
    try:
        dt_el = row.find_elements(By.CSS_SELECTOR, "td.date, span.date")
        if dt_el: info["신고 일시"] = _format_date(dt_el[0].text.strip())
    except: pass

    # 제목 (목록에서 미리 수집 - 2중 필터)
    try:
        subject_found = False
        subject_targets = row.find_elements(By.CSS_SELECTOR, "a.subject, .mail_subject, td.subject a, span.subject")
        for lk in subject_targets:
            txt = (lk.text or "").strip()
            if txt and len(txt) > 2 and not any(k in txt for k in ["첨부파일", "파일열기", "미리보기"]):
                raw_title = txt[:120]
                prefix_pattern = re.compile(r'^\[(신고메일|Report email|举报邮件)\]\s*', re.IGNORECASE)
                info["메일제목"] = prefix_pattern.sub('', raw_title)
                print(f"  [+] 원본 제목 수집 완료: {info['메일제목']}")
                subject_found = True
                break
        if not subject_found: # 폴백: td 텍스트 탐색
            for td in row.find_elements(By.TAG_NAME, "td"):
                txt = (td.text or "").strip()
                if txt and len(txt) > 5 and not any(k in txt for k in ["첨부파일", "요일", "오전", "오후"]):
                    if any(k in txt for k in ["[신고메일]", "[Report email]", "[举报邮件]"]):
                        prefix_pattern = re.compile(r'^\[(신고메일|Report email|举报邮件)\]\s*', re.IGNORECASE)
                        info["메일제목"] = prefix_pattern.sub('', txt[:120])
                        print(f"  [+] 목록 텍스트에서 제목 추출: {info['메일제목']}")
                        break
    except: pass

    # 메일 클릭
    try:
        clicked = False
        for target in row.find_elements(By.CSS_SELECTOR, "a.subject, td.subject a"):
            if target.is_displayed():
                driver.execute_script("arguments[0].click();", target)
                clicked = True; break
        if not clicked: driver.execute_script("arguments[0].click();", row)
        WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.ID, "readContentMessageWrap")))
    except: pass

    # 본문 데이터 파싱 (중국어 키워드 및 Fallback 포함)
    _extract_body_data(driver, info)

    # 팝업 처리 (미리보기 버튼 클릭)
    preview_clicked = False
    for sel in ['span[evt-rol="read-nested-pop"]', '.btn_fn4', '[evt-rol="read-nested-pop"]']:
        try:
            btns = driver.find_elements(By.CSS_SELECTOR, sel)
            for btn in btns:
                if btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    preview_clicked = True; time.sleep(1.5); break
            if preview_clicked: break
        except: continue
    
    if preview_clicked:
        _extract_from_popup(driver, info)
    else:
        _extract_from_mail_body_fallback(driver, info)
    
    return info

def _extract_body_data(driver, info):
    KW_MAP = [
        (["소속회사", "회사명", "그룹사", "소속", "affiliated company", "affiliatedcompany", "所属公司"], "그룹사"),
        (["부서명", "부서", "department", "部门"], "부서"),
        (["신고자", "성명", "이름", "담당자", "reporter's name", "reporter", "reportersname", "举报人姓名"], "신고자"),
    ]
    
    def _parse_table(ctx, require_class=True):
        res = {}
        containers = ctx.find_elements(By.CSS_SELECTOR, ".reportPhishing")
        if not containers and not require_class:
            containers = ctx.find_elements(By.CSS_SELECTOR, "table")
        for container in containers:
            for row in container.find_elements(By.TAG_NAME, "tr"):
                ths = row.find_elements(By.TAG_NAME, "th")
                tds = row.find_elements(By.TAG_NAME, "td")
                if not ths or not tds:
                    continue
                lbl = (ths[0].text or "").strip().replace(" ", "").lower()
                val = (tds[0].text or "").strip()
                for kws, field in KW_MAP:
                    if not res.get(field) and any(kw in lbl for kw in kws) and val:
                        res[field] = val[:100]
        return res

    parsed = {}

    # 1단계: 현재 컨텍스트 직접 파싱 (.reportPhishing 우선)
    if not any(parsed.values()):
        parsed = _parse_table(driver)
    # 1-1단계: .reportPhishing 없으면 일반 테이블도 시도
    if not any(parsed.values()):
        parsed = _parse_table(driver, require_class=False)

    # 2단계: 레벨1 iframe 탐색
    if not any(parsed.values()):
        lv1_iframes = driver.find_elements(By.CSS_SELECTOR, "#readContentMessageWrap iframe, iframe")
        for iframe1 in lv1_iframes:
            try:
                driver.switch_to.frame(iframe1)
                parsed = _parse_table(driver)
                if not any(parsed.values()):
                    parsed = _parse_table(driver, require_class=False)
                if any(parsed.values()):
                    driver.switch_to.default_content()
                    break

                # 3단계: 레벨2 중첩 iframe 탐색 (parent_frame으로 복귀)
                lv2_iframes = driver.find_elements(By.TAG_NAME, "iframe")
                for iframe2 in lv2_iframes:
                    try:
                        driver.switch_to.frame(iframe2)
                        parsed = _parse_table(driver)
                        if not any(parsed.values()):
                            parsed = _parse_table(driver, require_class=False)
                        driver.switch_to.parent_frame()  # 레벨1로 복귀
                        if any(parsed.values()):
                            break
                    except:
                        try: driver.switch_to.default_content()
                        except: pass
                        break

                driver.switch_to.default_content()
                if any(parsed.values()):
                    break
            except:
                try: driver.switch_to.default_content()
                except: pass

    for f, v in parsed.items(): 
        if v: info[f] = v

    # Fallback: 본문 전체 텍스트에서 정규식 추출
    if not all([info["그룹사"], info["부서"], info["신고자"]]):
        try:
            txt = driver.find_element(By.TAG_NAME, "body").text
            for field, pattern in [
                # 줄바꿈이 없는 부분, 띄어쓰기 외 특수문자 전까지만 자르는 [^\n\r<]+ 로 수정
                ("그룹사", re.compile(r'(?:소속회사|회사명|그룹사|소속|affiliated\s*company|所属公司)\s*[:：]\s*([^\n\r<]+)', re.I)),
                ("부서", re.compile(r'(?:부서명|부서|department|部门)\s*[:：]\s*([^\n\r<]+)', re.I)),
                ("신고자", re.compile(r'(?:신고자|성명|이름|담당자|reporter\'?s?\s*name|举报人姓名)[ \t]*[:：]?[ \t]*(?:이름|성함|이름은)?\s*([^\n\r<]+)', re.I))
            ]:
                if not info[field]:
                    m = pattern.search(txt)
                    if m: 
                        val = m.group(1).strip()
                        # "Reporter" 등의 경우를 대비해 email 관련 불필요 문자열 등 정제
                        val = re.split(r'(신고자|reporter|메일|mail)', val, flags=re.I)[0].strip()
                        if val: info[field] = val[:100]
        except: pass

def _extract_from_popup(driver, info):
    original_window = driver.current_window_handle
    time.sleep(2)
    all_windows = driver.window_handles
    if len(all_windows) > 1:
        new_win = [w for w in all_windows if w != original_window][-1]
        try:
            driver.switch_to.window(new_win)
            # 제목 (필터 적용)
            if not info["메일제목"]:
                for s in driver.find_elements(By.CSS_SELECTOR, "span.subject"):
                    t = s.text.strip()
                    if t and len(t) > 2 and not any(k in t for k in ["첨부파일", " 개", "KB", "MB"]):
                        info["메일제목"] = t; break
            
            # 발신/수신/날짜
            names = driver.find_elements(By.CSS_SELECTOR, "span.name")
            if len(names) >= 1: info["보낸사람"] = names[0].text.strip()
            if len(names) >= 2: info["받는사람"] = names[1].text.strip()
            dates = driver.find_elements(By.CSS_SELECTOR, "span.date")
            if dates: info["보낸날짜"] = _format_date(dates[0].text.strip())

            # 첨부파일 (필터 적용)
            body_txt = driver.find_element(By.TAG_NAME, "body").text
            att = []
            for m in re.finditer(r'([^\n\r]+?\.[a-zA-Z]{2,6})\s*\(\d[\d,\.]*\s*KB\)', body_txt, re.I):
                f = m.group(1).strip()
                if f and not any(k in f for k in ["단위", "MB", "GB"]): att.append(f)
            if att: info["첨부파일명"] = ", ".join(list(set(att)))

            driver.close()
        except: pass
        finally:
            driver.switch_to.window(original_window)

def _extract_from_mail_body_fallback(driver, info):
    # 팝업 실패 시 본문에서 직접 헤더 추출 로직 (셀렉터 기반)
    MAP = {"보낸사람": [".mail_from em", "#mailFrom"], "받는사람": [".mail_to em", "#mailTo"], "보낸날짜": [".mail_date em", "#mailDate"]}
    for field, sels in MAP.items():
        if not info[field]:
            for s in sels:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, s)
                    if el.is_displayed(): info[field] = el.text.strip(); break
                except: continue

def go_back_to_list(driver):
    try:
        # 1. '목록' 버튼 클릭
        for selector in ["button.btn_list", ".btn_list", "a[evt-rol='list']"]:
            for btn in driver.find_elements(By.CSS_SELECTOR, selector):
                if "목록" in (btn.text or "") and btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(1.5); return
        # 2. 뒤로가기
        driver.back(); time.sleep(2)
    except:
        if MALMAIL_URL: driver.get(MALMAIL_URL); time.sleep(2)


# ─────────────────────────────────────────────
# 엑셀 저장 및 메인
# ─────────────────────────────────────────────
def save_to_excel(records, base_dir):
    if not records: return
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(base_dir, f"malmail_info_{ts}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "악성메일 분석 목록"
    
    headers = ["No", "신고 일시", "메일 유형", "메일 유입 시간", "발신자", "수신자", "메일제목", "첨부파일명", "유형", "그룹사", "부서", "신고자"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    
    for ri, rec in enumerate(records, 2):
        row = [ri-1, rec.get("신고 일시"), "악성 메일", rec.get("보낸날짜"), rec.get("보낸사람"), rec.get("받는사람"), rec.get("메일제목"), rec.get("첨부파일명"), "유입 신고", rec.get("그룹사"), rec.get("부서"), rec.get("신고자")]
        for ci, val in enumerate(row, 1): ws.cell(row=ri, column=ci, value=val)
    
    wb.save(path)
    print(f"\n[SUCCESS] 엑셀 저장 완료: {path}")

def main():
    print("="*60 + "\n  ◈악성메일◈ 편지함 정보 추출 (최신 패치 v3.5)\n" + "="*60)
    
    # 사용자 입력 대기 (원상 복구)
    limit = input("  추출할 메일 건수 (전체: Enter): ").strip()
    target_limit = int(limit) if limit.isdigit() else None

    config = load_config()
    driver = login_shinhan_mail(config)
    if not driver: sys.exit(1)
    
    driver.implicitly_wait(0)
    try:
        if not navigate_to_malmail_folder(driver):
            print("[FAIL] 폴더 이동 실패"); return
        
        global MALMAIL_URL; MALMAIL_URL = driver.current_url
        mail_ids = collect_all_mail_ids(driver, target_limit)
        
        records = []
        for idx, mid in enumerate(mail_ids):
            print(f"\n[*] 메일 {idx+1}/{len(mail_ids)} (ID: {mid})")
            # 현재 페이지에 없으면 다음 페이지 탐색 로직 (생략 - 필요 시 추가)
            info = open_mail_and_extract_info(driver, mid)
            if info:
                # 발신자 이메일 정제
                m = re.search(r'<([^>]+)>', info["보낸사람"])
                if m: info["보낸사람"] = m.group(1).strip()
                records.append(info)
                print(f"  제목: {info['메일제목'][:30]}... / 발신: {info['보낸사람']}")
            
            go_back_to_list(driver)
        
        base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        save_to_excel(records, base_dir)
    finally:
        driver.quit()
        print("[*] 종료")

if __name__ == "__main__":
    main()
